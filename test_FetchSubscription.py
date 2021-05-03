import requests
import json
import jsonpath
import pytest
import openpyxl

def test_fetchsubscription(SubsId):
            #wb=openpyxl.load_workbook('//Users//rebeca.modrego//Desktop//API//Subscriptions.xlsx') #open the excel file
            #sh=wb['Sheet1']
            #rows=sh.max_row #how many rows are in the excel file
            #for i in range (2, rows+1):
                #cell_SubsId=sh.cell(row=i, column=1)
                #SubsId=cell_SubsId.value
                url="https://ysita-bssapi.staging.qvantel.net/v3/yoigo/revenue/subscriptions/" + str(SubsId)
                #Send GET request
                response=requests.get(url)
                #Display de BODY of the Output JSON
                json_response=json.loads(response.content)
                msisdn = jsonpath.jsonpath(json_response, 'data.attributes.msisdn')
                customerAcc = jsonpath.jsonpath(json_response, 'data.relationships.customer-account.data.id')
                return (msisdn[0], customerAcc[0], SubsId)



