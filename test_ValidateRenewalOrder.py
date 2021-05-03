import test_FetchSubscription
import test_RenewalQ25Extension
import test_VerifyOrderEligilibility
import test_VerifyFinancingConditions
import test_RenewalSIMOnly
import test_FetchBillingAccount
import pytest
import openpyxl


@pytest.mark.Smoke
def test_ValidateRenewal():
    #Excel (open)
    wb = openpyxl.load_workbook('//Users//rebeca.modrego//Desktop//API//Subscriptions.xlsx')  # open the excel file
    sh = wb['Sheet1'] #Go to Sheet1
    rows = sh.max_row  # how many rows are in Sheet1
    for i in range(2, rows + 1):
        cell_SubsId = sh.cell(row=i, column=1)
        SubsId = cell_SubsId.value
        s=test_FetchSubscription.test_fetchsubscription(SubsId)
        msisdn=s[0]
        custAcc=s[1]
        newSubsId=s[2]

        bi=test_FetchBillingAccount.test_fetch_BAcc(custAcc)
        billAcc=bi[0]
        #z=test_VerifyFinancingConditions.test_financingconditions(msisdn)
        y=test_VerifyOrderEligilibility.test_verifyordereligibility(msisdn, custAcc, SubsId)
        print ("Is renewal order allowed?: " + str(y[0]))
        if y[0]==True:
            z=test_VerifyFinancingConditions.test_financingconditions(msisdn)
            TERM = z[1]
            PAYTERM = z[2]
            discount=z[3]
            #print (PAYTERM)
            #print(TERM)
            if (z[0]=="allowed"):
                print("Q25 Extension is allowed")
                ext=test_RenewalQ25Extension.test_renewalQ25extension(custAcc, newSubsId, billAcc, int(discount), PAYTERM, TERM)
                print("Status Code:", ext[1])
                assert (ext[1]==201)

            elif (z[0]!="allowed"):
                print("Q25 Extension is not allowed")
            #SIM Only
                if str(TERM=="None") and str(PAYTERM)=="None":
                    # TERM, PAYTERM, no existen -> SIM only, NO TERMINAL
                    sim=test_RenewalSIMOnly.test_renewalsimonly(custAcc,newSubsId,billAcc)
                    print ("Status Code:", sim[1])
                    assert (sim[1] == 201)
                            # TERM, PAYTERM, existen -> Q25 WITHOUT ACTION
                            #if str(TERM == "None") and str(PAYTERM) != "None"
                                #Q25 WITHOUT ACTION
                            # if str(TERM != "None") and str(PAYTERM) == "None"
                                # Q25 WITHOUT ACTION
                            # if str(TERM != "None") and str(PAYTERM) != "None"
                                # Q25 WITHOUT ACTION

        elif y[0]!=True:
            print("Renewal order cannot be done: " + y[1])